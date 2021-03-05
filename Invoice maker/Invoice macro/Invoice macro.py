from openpyxl import Workbook
from openpyxl import load_workbook
import os

path = "./files"
file_list = os.listdir(path)
print(file_list)


results = []

for file_name_raw in file_list:

    file_name = "./files/" + file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    # ws = wb.active
    ws = wb['작업지시서(기본)']

    result = []

    # result.append(file_name_raw)
    result.append(ws['A9'].value)
    result.append(ws['C9'].value)
    result.append(ws['D9'].value)
    result.append(ws['E9'].value)
    result.append(ws['F9'].value)
    result.append(ws['G9'].value)
    result.append(ws['H9'].value)
    result.append(ws['I9'].value)
    result.append(ws['K9'].value)

    results.append(result)
    result = []

    result.append(ws['A12'].value)
    result.append(ws['C12'].value)
    result.append(ws['D12'].value)
    result.append(ws['E12'].value)
    # result.append(ws['F12'].value)
    # result.append(ws['G12'].value)
    # result.append(ws['H12'].value)
    # result.append(ws['I12'].value)
    # result.append(ws['K12'].value)

    results.append(result)
    result = []

    result.append(ws['A15'].value)
    result.append(ws['C15'].value)
    result.append(ws['D15'].value)
    result.append(ws['E15'].value)
    # result.append(ws['F15'].value)
    # result.append(ws['G15'].value)
    # result.append(ws['H15'].value)
    # result.append(ws['I15'].value)
    # result.append(ws['K15'].value)

    results.append(result)
    result = []

    result.append(ws['A18'].value)
    result.append(ws['C18'].value)
    result.append(ws['D18'].value)
    result.append(ws['E18'].value)
    # result.append(ws['F18'].value)
    # result.append(ws['G18'].value)
    # result.append(ws['H18'].value)
    # result.append(ws['I18'].value)
    # result.append(ws['K18'].value)

    results.append(result)

print(results)

wb = Workbook()
ws = wb.active

for i in results:
    ws.append(i)

ws.insert_cols(2)
ws.insert_cols(10)

wb.save("c:/Users./Latex./Desktop./택배./택배.xlsx")
